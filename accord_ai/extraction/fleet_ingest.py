"""Fleet roster ingestion — parse Excel / CSV into Pydantic Vehicle + Driver lists.

Port of accord_ai_v3/extraction/fleet_ingest.py with v4 schema field names:
  dob → date_of_birth (type date), full_name → first_name+last_name split,
  gvwr → gvw, stated_value → stated_amount (Decimal), radius → radius_of_travel,
  garaging_{zip,state,city} flat → garage_address: Address.

New in v4:
  - Input is bytes + filename (API upload path, not filesystem path).
  - Trailing-totals row rejection (rows whose first non-empty cell matches
    "Total" / "Subtotal" / "Grand Total" are skipped before entity construction).
  - MVR status normalisation to Literal enum values.
  - Use-type normalisation to Literal enum values.

Internal pipeline: _load_rows → _find_header → _classify_columns → _extract_rows.
"""
from __future__ import annotations

import csv as _csv
import io
import logging
import re
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any

from accord_ai.schema import Address, Driver, Vehicle

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

_MAX_SHEETS = 20
_MAX_ROWS_PER_SHEET = 100_000

# ---------------------------------------------------------------------------
# Column keyword bank  (v4 field names throughout)
# ---------------------------------------------------------------------------

_COLUMN_KEYWORDS: dict[str, list[str]] = {
    # Driver
    "first_name":       ["first name", "given name", "firstname", "fname"],
    "last_name":        ["last name", "surname", "family name", "lastname", "lname"],
    "full_name":        ["full name", "driver full name", "name of driver",
                         "driver name", "dr name", "driver"],
    "date_of_birth":    ["date of birth", "dob", "birthdate", "birth date", "born"],
    "license_state":    ["license state", "dl state", "state of license",
                         "licensing state", "lic state"],
    "license_class":    ["cdl class", "license class"],
    "license_number":   ["license number", "license no", "license #",
                         "drivers license", "driver license", "dl number",
                         "cdl number", "dl"],
    "years_experience": ["years of experience", "years licensed", "years experience",
                         "driving experience", "yrs exp", "experience", "years"],
    "hire_date":        ["hire date", "date hired", "employment date", "hired"],
    "sex":              ["sex", "gender", "m/f"],
    "marital_status":   ["marital status", "marital", "married"],
    "mvr_status":       ["mvr status", "driving record", "record status", "mvr"],
    # Vehicle
    "vin":              ["vin number", "vehicle identification", "vehicle id",
                         "serial number", "vin"],
    "year":             ["model year", "year", "yr.", "yr"],
    "make":             ["manufacturer", "mfg", "make"],
    "model":            ["vehicle model", "model"],
    "body_type":        ["body type", "body style", "vehicle type", "type"],
    "stated_amount":    ["stated amount", "stated value", "vehicle value",
                         "market value", "purchase price", "value", "cost", "acv"],
    "garaging_zip":     ["garaging zip", "garage zip", "parking zip", "garaging postal"],
    "garaging_state":   ["garaging state", "garage state", "parking state"],
    "garaging_city":    ["garaging city", "garage city", "parking city"],
    "use_type":         ["vehicle use", "use type", "usage", "use"],
    "radius_of_travel": ["radius of travel", "travel radius", "radius"],
    "gvw":              ["gross vehicle weight", "weight rating", "gross weight", "gvwr", "gvw"],
    "cargo_type":       ["cargo type", "what you haul", "commodity", "cargo"],
    # Catch-all — evaluated last; "name" matches anything with that word
    "_name_fallback":   ["name"],
}

_NAME_FALLBACK_FIELD = "full_name"

_DRIVER_FIELDS = {
    "full_name", "first_name", "last_name", "date_of_birth",
    "license_number", "license_state", "hire_date", "years_experience", "mvr_status",
}
_VEHICLE_FIELDS = {
    "vin", "year", "make", "model", "body_type", "stated_amount",
    "garaging_zip", "garaging_state", "garaging_city",
    "use_type", "radius_of_travel", "gvw", "cargo_type",
}
_GARAGING_COLS = {"garaging_zip", "garaging_state", "garaging_city"}

_VALUE_PATTERNS = {
    "vin":            re.compile(r"^[A-HJ-NPR-Z0-9]{17}$"),
    "year":           re.compile(r"^(19|20)\d{2}$"),
    "dob_slash":      re.compile(r"^\d{1,2}[/-]\d{1,2}[/-]\d{2,4}$"),
    "email":          re.compile(r"^[\w.+-]+@[\w.-]+\.[a-z]{2,}$", re.IGNORECASE),
    "phone":          re.compile(r"^\(?\d{3}\)?[\s.-]?\d{3}[\s.-]?\d{4}$"),
    "zip":            re.compile(r"^\d{5}(-\d{4})?$"),
    "state":          re.compile(r"^[A-Z]{2}$"),
    "two_words_name": re.compile(r"^[A-Z][a-z]+(\s+[A-Z][a-z]+)+$"),
}

_TOTAL_ROW_RE = re.compile(
    r"^\s*(grand\s+)?total\b|^\s*subtotal\b", re.IGNORECASE
)

_MVR_MAP: list[tuple[re.Pattern[str], str]] = [
    (re.compile(r"\bclean\b|\bclear\b|\bnone\b|\bno\s+violations?\b", re.I), "clean"),
    (re.compile(r"\bminor\b", re.I), "minor_violations"),
    (re.compile(r"\bmajor\b|\bdui\b|\bdwi\b|\bfelony\b", re.I), "major_violations"),
    (re.compile(r"\bsuspend|\brevoke", re.I), "suspended"),
]

_USE_TYPE_MAP: list[tuple[re.Pattern[str], str]] = [
    (re.compile(r"\bservice\b", re.I), "service"),
    (re.compile(r"\bcommercial\b|\bfor.hire\b|\bfreight\b", re.I), "commercial"),
    (re.compile(r"\bretail\b|\bdelivery\b", re.I), "retail"),
    (re.compile(r"\bpleasure\b|\bpersonal\b|\bprivate\b", re.I), "pleasure"),
]


# ---------------------------------------------------------------------------
# Public result type
# ---------------------------------------------------------------------------

@dataclass(frozen=True)
class FleetIngestResult:
    vehicles: list[Vehicle]
    drivers: list[Driver]
    header_row_idx: int           # 0-based index in raw rows; -1 = not found
    columns_classified: dict[int, str]   # col_index → canonical field name
    warnings: list[str]


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def parse_fleet_sheet(content: bytes, filename: str) -> FleetIngestResult:
    """Parse raw fleet file bytes into structured vehicles + drivers.

    Detects format from filename extension (.xlsx/.xls/.xlsm → openpyxl,
    anything else → CSV). Multi-sheet workbooks are processed sheet by sheet;
    header_row_idx and columns_classified reflect the first successfully
    parsed sheet.
    """
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext in ("xlsx", "xls", "xlsm"):
        return _parse_xlsx(content, filename)
    return _parse_csv(content, filename)


# ---------------------------------------------------------------------------
# Format loaders
# ---------------------------------------------------------------------------

def _parse_xlsx(content: bytes, filename: str) -> FleetIngestResult:
    import openpyxl

    warnings: list[str] = []
    vehicles: list[Vehicle] = []
    drivers: list[Driver] = []
    primary_header_idx = -1
    primary_columns: dict[int, str] = {}

    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(content), data_only=True, read_only=True
        )
    except MemoryError:
        return FleetIngestResult(
            vehicles=[], drivers=[], header_row_idx=-1,
            columns_classified={},
            warnings=["refused: workbook exceeded memory limits (zip-bomb?)"],
        )

    if len(wb.sheetnames) > _MAX_SHEETS:
        try:
            wb.close()
        except Exception:
            pass
        return FleetIngestResult(
            vehicles=[], drivers=[], header_row_idx=-1,
            columns_classified={},
            warnings=[
                f"refused: {len(wb.sheetnames)} sheets (max {_MAX_SHEETS}; "
                "possible zip-bomb)"
            ],
        )

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows: list[tuple[Any, ...]] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= _MAX_ROWS_PER_SHEET:
                warnings.append(
                    f"sheet {sheet_name!r} truncated at {_MAX_ROWS_PER_SHEET} rows"
                )
                break
            rows.append(row)
        if not rows:
            continue

        sheet_warns: list[str] = []
        hdr_idx, cols, v_list, d_list = _process_sheet(
            rows, sheet_name, sheet_warns
        )
        warnings.extend(sheet_warns)
        vehicles.extend(v_list)
        drivers.extend(d_list)

        if primary_header_idx == -1 and hdr_idx >= 0:
            primary_header_idx = hdr_idx
            primary_columns = cols

    try:
        wb.close()
    except Exception:
        pass

    return FleetIngestResult(
        vehicles=vehicles,
        drivers=drivers,
        header_row_idx=primary_header_idx,
        columns_classified=primary_columns,
        warnings=warnings,
    )


def _parse_csv(content: bytes, filename: str) -> FleetIngestResult:
    warnings: list[str] = []
    text = content.decode("utf-8-sig", errors="replace")
    reader = _csv.reader(io.StringIO(text))
    rows = [tuple(r) for r in reader]
    if not rows:
        return FleetIngestResult(
            vehicles=[], drivers=[], header_row_idx=-1,
            columns_classified={}, warnings=warnings,
        )

    sheet_name = filename.rsplit(".", 1)[0] if "." in filename else filename
    hdr_idx, cols, vehicles, drivers = _process_sheet(rows, sheet_name, warnings)
    return FleetIngestResult(
        vehicles=vehicles,
        drivers=drivers,
        header_row_idx=hdr_idx,
        columns_classified=cols,
        warnings=warnings,
    )


# ---------------------------------------------------------------------------
# Core pipeline
# ---------------------------------------------------------------------------

def _process_sheet(
    rows: list[tuple[Any, ...]],
    sheet_name: str,
    warnings: list[str],
) -> tuple[int, dict[int, str], list[Vehicle], list[Driver]]:
    """Run the full pipeline on one sheet's rows.

    Returns (header_row_idx, columns_classified, vehicles, drivers).
    """
    hdr_idx = _find_header(rows)
    if hdr_idx < 0:
        warnings.append(
            f"Sheet '{sheet_name}': no clear header row found "
            f"(skipped {len(rows)} rows)"
        )
        return -1, {}, [], []

    header_row = rows[hdr_idx]
    data_rows = rows[hdr_idx + 1:]

    cols = _classify_columns(header_row, data_rows)
    if not cols:
        warnings.append(
            f"Sheet '{sheet_name}': no classifiable columns found"
        )
        return hdr_idx, {}, [], []

    vehicles, drivers, row_warns = _extract_rows(data_rows, cols)
    for w in row_warns:
        warnings.append(f"Sheet '{sheet_name}': {w}")

    return hdr_idx, cols, vehicles, drivers


def _find_header(
    rows: list[tuple[Any, ...]], max_scan: int = 10
) -> int:
    """Return the 0-based row index that looks most like a column header.

    Scoring: keyword hits * 10 + string cells − numeric cells * 2.
    Returns -1 if no row scores >= 10 (no recognisable keyword match).
    """
    best_row = 0
    best_score = -1
    for i, row in enumerate(rows[:max_scan]):
        non_empty = [c for c in row if c is not None and str(c).strip()]
        if not non_empty:
            continue
        keyword_hits = string_cells = numeric_cells = 0
        for c in non_empty:
            if isinstance(c, (int, float)):
                numeric_cells += 1
                continue
            s = _norm_header(c)
            if not s:
                continue
            string_cells += 1
            for kw, _ in _KEYWORD_INDEX:
                if kw in s:
                    keyword_hits += 1
                    break
        score = keyword_hits * 10 + string_cells - numeric_cells * 2
        if score > best_score:
            best_score = score
            best_row = i
    return best_row if best_score >= 10 else -1


def _classify_columns(
    header_row: tuple[Any, ...],
    data_rows: list[tuple[Any, ...]],
) -> dict[int, str]:
    """Map column indices to canonical v4 field names.

    Strategy: keyword match on header cell (longest phrase wins), then
    value-shape pattern match on sample data as fallback.
    """
    col_samples: dict[int, list[Any]] = {}
    for row in data_rows[:30]:
        for i, v in enumerate(row):
            col_samples.setdefault(i, []).append(v)

    result: dict[int, str] = {}
    for i, h in enumerate(header_row):
        samples = col_samples.get(i, [])
        cf = _classify_one_column(h, samples)
        if cf:
            result[i] = cf
    return result


def _classify_one_column(header: Any, sample_values: list[Any]) -> str | None:
    h = _norm_header(header)
    if h:
        for kw, target_field in _KEYWORD_INDEX:
            if kw in h:
                return target_field

    non_empty = [_stringify(v) for v in sample_values if v is not None and str(v).strip()]
    if not non_empty:
        return None

    checks: list[tuple[str, re.Pattern[str], float]] = [
        ("vin",           _VALUE_PATTERNS["vin"],            0.8),
        ("year",          _VALUE_PATTERNS["year"],           0.8),
        ("date_of_birth", _VALUE_PATTERNS["dob_slash"],      0.7),
        ("email",         _VALUE_PATTERNS["email"],          0.8),
        ("phone",         _VALUE_PATTERNS["phone"],          0.8),
        ("garaging_zip",  _VALUE_PATTERNS["zip"],            0.6),
        ("license_state", _VALUE_PATTERNS["state"],          0.4),
        ("full_name",     _VALUE_PATTERNS["two_words_name"], 0.5),
    ]
    sample = non_empty[:20]
    for fname, pat, threshold in checks:
        hits = sum(1 for v in sample if pat.match(v))
        if hits / len(sample) >= threshold:
            return fname
    return None


def _extract_rows(
    data_rows: list[tuple[Any, ...]],
    cols: dict[int, str],
) -> tuple[list[Vehicle], list[Driver], list[str]]:
    """Build Vehicle and Driver instances from data rows.

    Trailing-totals rows (first non-empty cell matches "Total"/"Subtotal")
    are rejected before entity construction.
    """
    driver_cols = {i: f for i, f in cols.items() if f in _DRIVER_FIELDS}
    vehicle_cols = {i: f for i, f in cols.items() if f in _VEHICLE_FIELDS}

    sheet_type = _classify_sheet_type(list(cols.values()))

    vehicles: list[Vehicle] = []
    drivers: list[Driver] = []
    warnings: list[str] = []

    for row in data_rows:
        # Skip entirely blank rows
        if not any(c is not None and str(c).strip() for c in row):
            continue

        # Trailing-totals rejection
        first_str = next(
            (str(c).strip() for c in row if c is not None and str(c).strip()), ""
        )
        if _TOTAL_ROW_RE.match(first_str):
            continue

        if driver_cols and sheet_type in ("drivers", "mixed"):
            raw = _row_to_raw(row, driver_cols)
            if raw:
                d = _build_driver(raw)
                if d is not None:
                    drivers.append(d)

        if vehicle_cols and sheet_type in ("vehicles", "mixed"):
            raw = _row_to_raw(row, vehicle_cols)
            if raw:
                v = _build_vehicle(raw, warnings)
                if v is not None:
                    vehicles.append(v)

    return vehicles, drivers, warnings


# ---------------------------------------------------------------------------
# Entity builders
# ---------------------------------------------------------------------------

def _build_driver(raw: dict[str, Any]) -> Driver | None:
    """Construct a Driver from a cleaned raw-field dict."""
    # Compose first/last from full_name if needed
    if "full_name" in raw and "first_name" not in raw and "last_name" not in raw:
        parts = str(raw.pop("full_name")).split(None, 1)
        raw["first_name"] = parts[0]
        if len(parts) > 1:
            raw["last_name"] = parts[1]
    elif "full_name" in raw:
        raw.pop("full_name")  # first/last already present; drop composite

    # Must have at minimum a name fragment or license
    has_name = raw.get("first_name") or raw.get("last_name")
    has_license = raw.get("license_number")
    if not has_name and not has_license:
        return None

    kwargs: dict[str, Any] = {}
    for field_name, val in raw.items():
        cleaned = _clean_value(field_name, val)
        if cleaned is not None:
            kwargs[field_name] = cleaned
    try:
        return Driver(**kwargs)
    except Exception as exc:
        logger.debug("Driver construction failed: %s  raw=%r", exc, raw)
        return None


def _build_vehicle(raw: dict[str, Any], warnings: list[str]) -> Vehicle | None:
    """Construct a Vehicle from a cleaned raw-field dict.

    Collapses garaging_{zip,state,city} into garage_address: Address.
    """
    # Must have VIN or year+make+model
    has_vin = bool(raw.get("vin"))
    has_ymm = raw.get("year") and raw.get("make") and raw.get("model")
    if not has_vin and not has_ymm:
        return None

    kwargs: dict[str, Any] = {}
    garaging: dict[str, Any] = {}

    for field_name, val in raw.items():
        if field_name in _GARAGING_COLS:
            cleaned = _clean_value(field_name, val)
            if cleaned is not None:
                addr_key = field_name.replace("garaging_", "")  # zip→zip_code below
                if addr_key == "zip":
                    garaging["zip_code"] = cleaned
                else:
                    garaging[addr_key] = cleaned
        else:
            cleaned = _clean_value(field_name, val)
            if cleaned is not None:
                kwargs[field_name] = cleaned

    if garaging:
        kwargs["garage_address"] = Address(**garaging)

    if "vin" in kwargs:
        vin = str(kwargs["vin"]).replace(" ", "").upper()
        if re.match(r"^\d+\.\d+E\+\d+$", vin, re.IGNORECASE):
            warnings.append(
                f"VIN {vin!r} appears to be Excel scientific-notation corruption; "
                "needs manual correction"
            )
            del kwargs["vin"]

    try:
        return Vehicle(**kwargs)
    except Exception as exc:
        logger.debug("Vehicle construction failed: %s  raw=%r", exc, raw)
        return None


# ---------------------------------------------------------------------------
# Value cleaner
# ---------------------------------------------------------------------------

def _clean_value(field_name: str, raw: Any) -> Any:  # noqa: PLR0911
    """Coerce a raw cell value to the type expected by the v4 schema field."""
    if raw is None:
        return None
    if isinstance(raw, str) and not raw.strip():
        return None

    s = _stringify(raw)

    if field_name in ("date_of_birth", "hire_date"):
        return _coerce_date(raw)

    if field_name == "vin":
        v = s.replace(" ", "").upper()
        return v or None

    if field_name == "year":
        try:
            y = int(float(s))
            return y if 1900 <= y <= 2100 else None
        except (ValueError, TypeError):
            return None

    if field_name in ("stated_amount",):
        cleaned = re.sub(r"[^\d.]", "", s)
        if not cleaned:
            return None
        try:
            return Decimal(cleaned)
        except InvalidOperation:
            return None

    if field_name in ("gvw", "radius_of_travel", "years_experience"):
        cleaned = re.sub(r"[^\d.]", "", s)
        if not cleaned:
            return None
        try:
            return int(float(cleaned))
        except (ValueError, TypeError):
            return None

    if field_name == "license_state":
        v = s.upper()
        return v if _VALUE_PATTERNS["state"].match(v) else v[:2].upper() if len(v) >= 2 else None

    if field_name == "sex":
        v = s.upper()[:1]
        return v if v in ("M", "F") else None

    if field_name == "marital_status":
        v = s.upper()[:1]
        return v if v in ("S", "M", "D", "W", "P") else None

    if field_name == "mvr_status":
        for pat, canonical in _MVR_MAP:
            if pat.search(s):
                return canonical
        return None

    if field_name == "use_type":
        for pat, canonical in _USE_TYPE_MAP:
            if pat.search(s):
                return canonical
        return None

    if field_name in ("first_name", "last_name", "full_name"):
        v = re.sub(r"\s+", " ", s).strip()
        if v.isupper():
            v = v.title()
        return v or None

    return _defuse_formula(s)


# ---------------------------------------------------------------------------
# Sheet-type classifier
# ---------------------------------------------------------------------------

def _classify_sheet_type(column_fields: list[str]) -> str:
    present = set(f for f in column_fields if f)
    driver_hits = len(present & _DRIVER_FIELDS)
    vehicle_hits = len(present & _VEHICLE_FIELDS)
    if driver_hits >= 2 and vehicle_hits == 0:
        return "drivers"
    if vehicle_hits >= 2 and driver_hits == 0:
        return "vehicles"
    if driver_hits >= 2 and vehicle_hits >= 2:
        return "mixed"
    if driver_hits >= 1:
        return "drivers"
    if vehicle_hits >= 1:
        return "vehicles"
    return "unknown"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _build_keyword_index() -> list[tuple[str, str]]:
    pairs: list[tuple[str, str]] = []
    for field_name, keywords in _COLUMN_KEYWORDS.items():
        target = _NAME_FALLBACK_FIELD if field_name == "_name_fallback" else field_name
        for kw in keywords:
            pairs.append((kw, target))
    pairs.sort(key=lambda p: -len(p[0]))
    return pairs


_KEYWORD_INDEX = _build_keyword_index()


def _norm_header(h: Any) -> str:
    if h is None:
        return ""
    s = re.sub(r"[^\w\s/#.-]", "", str(h).strip().lower())
    return re.sub(r"\s+", " ", s).strip()


def _stringify(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, (date, datetime)):
        return v.strftime("%m/%d/%Y")
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _defuse_formula(s: str) -> str:
    if isinstance(s, str) and s and s[0] in ("=", "+", "-", "@"):
        return "'" + s
    return s


def _coerce_date(raw: Any) -> date | None:
    """Return a date object, or None on failure."""
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    if isinstance(raw, (int, float)):
        # Excel serial date (1900 epoch)
        try:
            from datetime import timedelta
            base = datetime(1899, 12, 30)
            return (base + timedelta(days=int(raw))).date()
        except Exception:
            return None
    s = str(raw).strip()
    if not s:
        return None
    for fmt in (
        "%m/%d/%Y", "%m-%d-%Y", "%Y-%m-%d", "%Y/%m/%d",
        "%m/%d/%y", "%d/%m/%Y", "%b %d, %Y", "%d %b %Y",
    ):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _row_to_raw(row: tuple[Any, ...], fields_idx: dict[int, str]) -> dict[str, Any]:
    """Pluck the relevant columns from a row into a raw field dict."""
    out: dict[str, Any] = {}
    for i, field_name in fields_idx.items():
        if i < len(row) and row[i] is not None and str(row[i]).strip():
            out[field_name] = row[i]
    return out


# ---------------------------------------------------------------------------
# Session merge (called by the endpoint layer)
# ---------------------------------------------------------------------------

def merge_fleet_into_submission(
    submission_dict: dict[str, Any],
    fleet: FleetIngestResult,
) -> dict[str, int]:
    """Merge fleet vehicles+drivers into a CustomerSubmission dict in-place.

    LOB rules (v4-new):
      - If lob_details is absent → set to commercial_auto.
      - If lob_details.lob != commercial_auto → raise ValueError (caller → 422).

    Dedup delegates to core/vehicle_merge.py (3-tier VIN-primary for vehicles,
    license-primary for drivers) — the single merge authority for all paths.

    Returns counts: {drivers_added, drivers_updated, vehicles_added, vehicles_updated}.
    """
    from accord_ai.core.vehicle_merge import merge_drivers, merge_vehicles
    from accord_ai.schema import Driver, Vehicle

    lob = submission_dict.get("lob_details", {})
    if not lob:
        submission_dict["lob_details"] = {"lob": "commercial_auto"}
        lob = submission_dict["lob_details"]
    elif lob.get("lob") != "commercial_auto":
        raise ValueError(
            f"Fleet upload requires commercial_auto LOB; "
            f"submission has lob={lob.get('lob')!r}"
        )

    lob.setdefault("drivers", [])
    lob.setdefault("vehicles", [])

    # Convert existing dicts to Pydantic so merge_vehicles / merge_drivers
    # operate on typed models regardless of the call path.
    existing_v = [Vehicle.model_validate(v) for v in lob["vehicles"]]
    existing_d = [Driver.model_validate(d) for d in lob["drivers"]]

    merged_v = merge_vehicles(existing_v, fleet.vehicles)
    merged_d = merge_drivers(existing_d, fleet.drivers)

    lob["vehicles"] = [v.model_dump(exclude_none=True) for v in merged_v]
    lob["drivers"] = [d.model_dump(exclude_none=True) for d in merged_d]

    vehicles_added = len(merged_v) - len(existing_v)
    drivers_added = len(merged_d) - len(existing_d)

    return {
        "vehicles_added":   vehicles_added,
        "vehicles_updated": len(fleet.vehicles) - vehicles_added,
        "drivers_added":    drivers_added,
        "drivers_updated":  len(fleet.drivers) - drivers_added,
    }
